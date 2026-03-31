"""
Generador de reportes Excel para Caja Mágica.
Paleta Emerald Obsidian: verde bosque sobre fondos claros,
celdas editables en azul, totales en amarillo suave.
"""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ─── Paleta Caja Mágica ─────────────────────────────────────────────────────

COLOR_HEADER_BG   = "1A4731"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ROW_ALT     = "F0F7F0"
COLOR_ROW_PLAIN   = "FFFFFF"
COLOR_TOTAL_BG    = "FFF9C4"
COLOR_TOTAL_FONT  = "1A4731"
COLOR_BORDER      = "CCCCCC"
COLOR_INPUT_BG    = "EBF5FB"
COLOR_INPUT_FONT  = "1F4E79"
FONT_NAME         = "Trebuchet MS"

COP_FMT = '#,##0" COP"'
USD_FMT = '"$"#,##0.00'

_BORDER_THIN = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)


def _apply_header(ws, row: int, cols: int) -> None:
    """Aplica estilo de encabezado a una fila."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FONT, size=11)
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER_THIN


def _apply_data_row(ws, row: int, cols: int, alt: bool = False) -> None:
    """Aplica estilo de fila de datos (alternando fondos)."""
    bg = COLOR_ROW_ALT if alt else COLOR_ROW_PLAIN
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_NAME, size=10)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        cell.border = _BORDER_THIN


def _apply_total_row(ws, row: int, cols: int) -> None:
    """Aplica estilo de fila totalizadora."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_NAME, bold=True, color=COLOR_TOTAL_FONT, size=11)
        cell.fill = PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_THIN


def _apply_input_row(ws, row: int, cols: int) -> None:
    """Aplica estilo de celda editable (azul claro)."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_NAME, color=COLOR_INPUT_FONT, size=10)
        cell.fill = PatternFill(start_color=COLOR_INPUT_BG, end_color=COLOR_INPUT_BG, fill_type="solid")
        cell.alignment = Alignment(vertical="center")
        cell.border = _BORDER_THIN


def _auto_col_width(ws, min_w: int = 10, max_w: int = 45) -> None:
    """Ajusta ancho de columnas basado en contenido."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 3, min_w), max_w)
        ws.column_dimensions[col_letter].width = adjusted


def generar_excel(movimientos: list, mes: str) -> str:
    """
    Genera archivo Excel con 4 hojas y paleta Caja Mágica.

    Args:
        movimientos: Lista de dicts de movimientos filtrados por mes.
        mes: String del mes en formato YYYY-MM.

    Returns:
        Ruta absoluta del archivo generado.
    """
    wb = Workbook()
    fecha_export = datetime.now().strftime("%Y-%m-%d %H:%M")

    # ═══════════════════ HOJA 1: Registro ═══════════════════════════════════

    ws1 = wb.active
    ws1.title = "Registro"

    # Título
    ws1.merge_cells("A1:I1")
    ws1["A1"] = f"CAJA MÁGICA — Registro {mes}"
    ws1["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_HEADER_FONT)
    ws1["A1"].fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Subtítulo
    ws1.merge_cells("A2:I2")
    ws1["A2"] = f"Caja Mágica · Tu tesorería personal · Exportado: {fecha_export}"
    ws1["A2"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Fecha", "Hora", "Descripción", "Tipo", "Categoría", "Moneda", "Monto Original", "Monto COP", "Proyección"]
    for col_i, h in enumerate(headers, 1):
        ws1.cell(row=3, column=col_i, value=h)
    _apply_header(ws1, 3, len(headers))

    # Datos ordenados por timestamp ASC
    sorted_movs = sorted(movimientos, key=lambda m: m.get("timestamp", ""))
    data_start = 4
    for i, mov in enumerate(sorted_movs):
        row = data_start + i
        ts = mov.get("timestamp", "")
        fecha = ts[:10] if len(ts) >= 10 else ""
        hora = ts[11:16] if len(ts) >= 16 else ""

        ws1.cell(row=row, column=1, value=fecha)
        ws1.cell(row=row, column=2, value=hora)
        ws1.cell(row=row, column=3, value=mov.get("descripcion", ""))
        ws1.cell(row=row, column=4, value=mov.get("tipo", ""))
        ws1.cell(row=row, column=5, value=mov.get("categoria", ""))
        ws1.cell(row=row, column=6, value=mov.get("moneda", "COP"))

        monto_orig = mov.get("monto_original", 0)
        moneda = mov.get("moneda", "COP")
        cell_orig = ws1.cell(row=row, column=7, value=monto_orig)
        cell_orig.number_format = USD_FMT if moneda == "USD" else COP_FMT

        cell_cop = ws1.cell(row=row, column=8, value=mov.get("monto_cop", 0))
        cell_cop.number_format = COP_FMT

        ws1.cell(row=row, column=9, value="Sí" if mov.get("es_proyeccion") else "No")

        _apply_data_row(ws1, row, len(headers), alt=(i % 2 == 1))

    # Fila TOTAL
    total_row = data_start + len(sorted_movs)
    ws1.cell(row=total_row, column=1, value="TOTAL")
    col_cop_letter = get_column_letter(8)
    ws1.cell(
        row=total_row,
        column=8,
        value=f"=SUM({col_cop_letter}{data_start}:{col_cop_letter}{total_row - 1})" if sorted_movs else 0,
    )
    ws1.cell(row=total_row, column=8).number_format = COP_FMT
    _apply_total_row(ws1, total_row, len(headers))

    _auto_col_width(ws1)

    # ═══════════════════ HOJA 2: Resumen ════════════════════════════════════

    ws2 = wb.create_sheet("Resumen")
    ws2.merge_cells("A1:B1")
    ws2["A1"] = f"RESUMEN — {mes}"
    ws2["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_HEADER_FONT)
    ws2["A1"].fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    ws2["A1"].alignment = Alignment(horizontal="center")

    metricas = [
        ("Mes analizado", mes),
        ("Total ingresos COP", f'=SUMPRODUCT((Registro!D{data_start}:D{total_row - 1}="ingreso")*Registro!H{data_start}:H{total_row - 1})'),
        ("Total egresos COP", f'=SUMPRODUCT((Registro!D{data_start}:D{total_row - 1}="egreso")*Registro!H{data_start}:H{total_row - 1})'),
        ("Total ahorros COP", f'=SUMPRODUCT((Registro!D{data_start}:D{total_row - 1}="ahorro")*Registro!H{data_start}:H{total_row - 1})'),
        ("Flujo neto COP", "=B3-B4-B5"),
        ("Movimientos totales", f'=COUNTA(Registro!C{data_start}:C{total_row - 1})'),
    ]

    for i, (label, value) in enumerate(metricas):
        row = i + 2
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=1).font = Font(name=FONT_NAME, bold=True, size=11)
        ws2.cell(row=row, column=1).border = _BORDER_THIN

        cell_val = ws2.cell(row=row, column=2, value=value)
        cell_val.border = _BORDER_THIN
        if i in (1, 2, 3, 4):
            cell_val.number_format = COP_FMT
        cell_val.font = Font(name=FONT_NAME, size=11)
        if i == 0:
            _apply_header(ws2, row, 2)
            ws2.cell(row=row, column=1).font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FONT, size=11)

    _auto_col_width(ws2)

    # ═══════════════════ HOJA 3: Flujo 6 Meses ═════════════════════════════

    ws3 = wb.create_sheet("Flujo 6 Meses")
    ws3.merge_cells("A1:E1")
    ws3["A1"] = "PROYECCIÓN DE FLUJO — 6 MESES"
    ws3["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_HEADER_FONT)
    ws3["A1"].fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    ws3["A1"].alignment = Alignment(horizontal="center")

    flow_headers = ["Mes", "Ingresos Proyectados", "Egresos Proyectados", "Neto", "Caja Acumulada"]
    for col_i, h in enumerate(flow_headers, 1):
        ws3.cell(row=2, column=col_i, value=h)
    _apply_header(ws3, 2, len(flow_headers))

    # Mes 0: valores reales de Resumen
    ws3.cell(row=3, column=1, value=mes)
    ws3.cell(row=3, column=2, value="=Resumen!B3")
    ws3.cell(row=3, column=3, value="=Resumen!B4")
    ws3.cell(row=3, column=4, value="=B3-C3")
    ws3.cell(row=3, column=5, value="=D3")
    _apply_data_row(ws3, 3, 5)

    # Meses 1-5: editables
    year, month_num = int(mes[:4]), int(mes[5:7])
    for offset in range(1, 6):
        row = 3 + offset
        m = month_num + offset
        y = year + (m - 1) // 12
        m = ((m - 1) % 12) + 1
        ws3.cell(row=row, column=1, value=f"{y}-{m:02d}")
        ws3.cell(row=row, column=2, value=0)
        ws3.cell(row=row, column=3, value=0)
        ws3.cell(row=row, column=4, value=f"=B{row}-C{row}")
        ws3.cell(row=row, column=5, value=f"=E{row - 1}+D{row}")
        _apply_input_row(ws3, row, 5)
        # Keep formula cells non-input styled
        ws3.cell(row=row, column=4).font = Font(name=FONT_NAME, size=10)
        ws3.cell(row=row, column=5).font = Font(name=FONT_NAME, size=10)

    for col_i in range(2, 6):
        for row in range(3, 9):
            ws3.cell(row=row, column=col_i).number_format = COP_FMT

    # Nota al pie
    ws3.merge_cells("A10:E10")
    ws3["A10"] = "Las celdas en azul claro son editables — ingresa tus estimaciones"
    ws3["A10"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")

    _auto_col_width(ws3)

    # ═══════════════════ HOJA 4: Supuestos ══════════════════════════════════

    ws4 = wb.create_sheet("Supuestos")
    ws4.merge_cells("A1:B1")
    ws4["A1"] = "SUPUESTOS Y PARÁMETROS"
    ws4["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_HEADER_FONT)
    ws4["A1"].fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    ws4["A1"].alignment = Alignment(horizontal="center")

    supuestos = [
        ("Parámetro", "Valor"),
        ("Tasa COP/USD", 4200),
        ("Precio hora servicios (USD)", 25),
        ("Horas facturables/semana", 10),
        ("Caja mínima COP (umbral de alerta)", 800000),
        ("% reinversión de ingresos", 50),
        ("Objetivo ingreso mensual USD", 2000),
    ]

    for i, (label, value) in enumerate(supuestos):
        row = i + 2
        ws4.cell(row=row, column=1, value=label)
        ws4.cell(row=row, column=2, value=value)
        if i == 0:
            _apply_header(ws4, row, 2)
        else:
            _apply_input_row(ws4, row, 2)
            ws4.cell(row=row, column=1).font = Font(name=FONT_NAME, bold=True, size=10, color=COLOR_INPUT_FONT)

    if isinstance(supuestos[4][1], int):
        ws4.cell(row=6, column=2).number_format = COP_FMT

    _auto_col_width(ws4)

    # ═══════════════════ Guardar ════════════════════════════════════════════

    base_dir = Path(__file__).parent
    output_dir = base_dir / "data"
    output_dir.mkdir(exist_ok=True)
    filepath = output_dir / f"CajaMagica_{mes}.xlsx"
    wb.save(str(filepath))

    return str(filepath)
